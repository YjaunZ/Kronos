import pandas as pd
import numpy as np
import sys
import os
from datetime import datetime, timedelta
import holidays
import baostock as bs
import matplotlib
import platform
import time
import json
from pathlib import Path
from functools import wraps
from tqdm import tqdm  # 导入进度条库
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import DataPoint

# 根据操作系统选择合适后端
system = platform.system()
if system == "Darwin":  # macOS
    matplotlib.use('MacOSX')
elif system == "Windows":
    matplotlib.use('TkAgg')
else:  # Linux
    matplotlib.use('TkAgg')

# 添加项目路径
sys.path.append(os.path.join(os.path.dirname(__file__), ".."))

from model import Kronos, KronosTokenizer, KronosPredictor


def retry_on_failure(max_retries=3):
    """重试装饰器"""

    def decorator(func):
        @wraps(func)
        def wrapper(*args, **kwargs):
            for attempt in range(max_retries):
                try:
                    return func(*args, **kwargs)
                except Exception as e:
                    if attempt == max_retries - 1:
                        print(f"重试{max_retries}次后仍失败: {e}")
                        raise e
                    wait_time = 2 ** attempt  # 指数退避
                    print(f"请求失败，{wait_time}秒后重试: {e}")
                    time.sleep(wait_time)
            return None

        return wrapper

    return decorator


@retry_on_failure(max_retries=2)
def get_stock_data_baostock(symbol, days=500):
    """获取指定天数的股票数据（使用baostock）"""
    # 重新登录
    lg = bs.login()
    if lg.error_code != '0':
        raise Exception(f"登录失败: {lg.error_msg}")

    try:
        # 处理股票代码格式
        if symbol.startswith(('sh', 'sz')):
            code = symbol  # 已经是baostock格式
        elif symbol.startswith('6'):  # 上交所
            code = f"sh.{symbol}"
        elif symbol.startswith(('0', '3')):  # 深交所
            code = f"sz.{symbol}"
        else:
            code = f"sz.{symbol}"  # 默认深市

        # 添加延时避免请求过频
        time.sleep(0.5)

        # 使用baostock获取股票数据
        rs = bs.query_history_k_data_plus(
            code,
            "date,open,high,low,close,volume",
            start_date=(datetime.now() - timedelta(days=days)).strftime('%Y-%m-%d'),
            end_date=datetime.now().strftime('%Y-%m-%d'),
            frequency="d",
            adjustflag="3"  # 前复权
        )

        data_list = []
        while (rs.error_code == '0') & rs.next():
            data_list.append(rs.get_row_data())

        if not data_list:
            raise ValueError(f"无法获取股票 {symbol} 的数据")

        # 转换为DataFrame
        result = pd.DataFrame(data_list, columns=rs.fields)

        # 添加成交额字段（volume * close 的近似值）
        result['amount'] = pd.to_numeric(result['volume'], errors='coerce') * pd.to_numeric(result['close'],
                                                                                            errors='coerce')

        # 重命名列
        column_mapping = {
            'date': 'timestamps',
            'open': 'open',
            'high': 'high',
            'low': 'low',
            'close': 'close',
            'volume': 'volume',
            'amount': 'amount'
        }
        required_data = result.rename(columns=column_mapping)[list(column_mapping.values())].copy()

        # 转换时间戳格式
        required_data['timestamps'] = pd.to_datetime(required_data['timestamps'])

        # 确保数值列是数字类型
        numeric_columns = ['open', 'high', 'low', 'close', 'volume', 'amount']
        for col in numeric_columns:
            required_data[col] = pd.to_numeric(required_data[col], errors='coerce')

        # 删除包含NaN的行
        required_data = required_data.dropna()

        return required_data
    finally:
        # 确保登出
        bs.logout()


def get_stock_data_cached(symbol, days=500, cache_dir="./cache"):
    """
    获取指定天数的股票数据，带缓存机制
    """
    # 创建缓存目录
    cache_path = Path(cache_dir)
    cache_path.mkdir(exist_ok=True)

    # 生成缓存文件名
    cache_file = cache_path / f"{symbol}_{days}_days.json"

    # 检查缓存是否存在且未过期（24小时）
    if cache_file.exists():
        cache_time = datetime.fromtimestamp(cache_file.stat().st_mtime)
        if (datetime.now() - cache_time).total_seconds() < 24 * 3600:  # 24小时内有效
            try:
                with open(cache_file, 'r', encoding='utf-8') as f:
                    cached_data = json.load(f)

                # 转换回DataFrame
                df = pd.DataFrame(cached_data)
                df['timestamps'] = pd.to_datetime(df['timestamps'])
                numeric_columns = ['open', 'high', 'low', 'close', 'volume', 'amount']
                for col in numeric_columns:
                    df[col] = pd.to_numeric(df[col], errors='coerce')

                print(f"从缓存加载 {symbol} 数据")
                return df
            except Exception as e:
                print(f"读取缓存失败: {e}")

    # 缓存不存在或已过期，重新获取数据
    try:
        required_data = get_stock_data_baostock(symbol, days)
    except Exception as e:
        print(f"获取股票 {symbol} 数据失败: {e}")
        return None

    # 保存到缓存
    try:
        cache_dict = required_data.copy()
        cache_dict['timestamps'] = cache_dict['timestamps'].dt.strftime('%Y-%m-%d')
        with open(cache_file, 'w', encoding='utf-8') as f:
            json.dump(cache_dict.to_dict(orient='records'), f, ensure_ascii=False, indent=2)
    except Exception as e:
        print(f"保存缓存失败: {e}")

    return required_data


def generate_future_trading_days(start_date, num_days):
    """
    生成未来交易日时间戳（排除周末和节假日）
    """
    # 将起始日期转换为pandas Timestamp
    if isinstance(start_date, str):
        start_date = pd.to_datetime(start_date)
    elif isinstance(start_date, datetime):
        start_date = pd.to_datetime(start_date)

    # 获取中国法定节假日
    cn_holidays = holidays.China()

    trading_days = []
    current_date = start_date + timedelta(days=1)  # 从第二天开始

    while len(trading_days) < num_days:
        # 检查是否为工作日（周一到周五）且不是节假日
        if current_date.weekday() < 5 and current_date.date() not in cn_holidays:
            trading_days.append(current_date)

        current_date += timedelta(days=1)

    return trading_days


def get_stock_industry_mapping():
    """
    获取股票所属行业映射表
    """
    # 重新登录
    lg = bs.login()
    if lg.error_code != '0':
        print(f"登录baostock失败: {lg.error_msg}")
        bs.logout()
        return None

    try:
        # 查询股票行业分类信息
        rs = bs.query_stock_industry()
        industry_list = []
        while (rs.error_code == '0') & rs.next():
            industry_list.append(rs.get_row_data())

        if not industry_list:
            print("未获取到行业分类数据")
            return None

        # 转换为DataFrame
        industry_df = pd.DataFrame(industry_list, columns=rs.fields)

        # 确保字段存在
        if 'code' in industry_df.columns and 'industry' in industry_df.columns:
            result = pd.DataFrame({
                'code': industry_df['code'],
                'industry': industry_df['industry']
            })
            return result
        else:
            print("行业数据字段结构不符合预期")
            return None

    except Exception as e:
        print(f"获取股票行业映射失败: {e}")
        return None
    finally:
        bs.logout()


def select_representative_stocks_for_industry(stocks, stock_data_dict, n=20):
    """
    为行业选择代表性股票（基于成交量和成交额）
    """
    # 计算每只股票的平均成交量和成交额
    stock_metrics = []
    for stock in stocks:
        if stock in stock_data_dict and not stock_data_dict[stock].empty:
            # 计算平均成交量
            avg_volume = stock_data_dict[stock]['volume'].mean()
            # 计算平均成交额
            avg_amount = stock_data_dict[stock]['amount'].mean()
            # 综合指标：成交量 * 0.6 + 成交额 * 0.4
            composite_metric = avg_volume * 0.6 + avg_amount * 0.4
            stock_metrics.append((stock, composite_metric))

    # 按综合指标降序排列，选择前n只股票
    sorted_stocks = sorted(stock_metrics, key=lambda x: x[1], reverse=True)
    representative_stocks = [stock[0] for stock in sorted_stocks[:n]]

    return representative_stocks


def calculate_industry_index_from_components(stock_data_dict, industry_mapping):
    """
    基于成分股数据计算行业指数
    """
    # 按行业分组
    industry_groups = industry_mapping.groupby('industry')['code'].apply(list)

    industry_indices = {}

    for industry, stocks in tqdm(industry_groups.items(), desc="计算行业指数", unit="行业"):
        # 选择代表性股票（前20只）
        representative_stocks = select_representative_stocks_for_industry(stocks, stock_data_dict, n=20)

        # 获取该行业代表性股票的数据
        valid_stocks_data = []
        for stock in representative_stocks:
            if stock in stock_data_dict and stock_data_dict[stock] is not None and not stock_data_dict[stock].empty:
                valid_stocks_data.append(stock_data_dict[stock])

        if not valid_stocks_data:
            continue

        # 找到共同的日期范围
        common_dates = set(valid_stocks_data[0].index)
        for stock_data in valid_stocks_data[1:]:
            common_dates = common_dates.intersection(set(stock_data.index))

        if not common_dates:
            continue

        common_dates = sorted(list(common_dates))

        # 计算每日行业指数
        industry_open = []
        industry_high = []
        industry_low = []
        industry_close = []
        industry_volume = []
        industry_amount = []

        for date in common_dates:
            # 获取当日所有成分股数据
            day_data = []
            for stock_data in valid_stocks_data:
                if date in stock_data.index:
                    day_data.append(stock_data.loc[date])

            if not day_data:
                continue

            # 将当日数据转换为DataFrame
            day_df = pd.DataFrame(day_data)

            # 使用等权重计算行业指数
            avg_open = day_df['open'].mean()
            avg_high = day_df['high'].mean()
            avg_low = day_df['low'].mean()
            avg_close = day_df['close'].mean()
            avg_volume = day_df['volume'].sum()  # 成交量相加
            avg_amount = day_df['amount'].sum()  # 成交额相加

            industry_open.append(avg_open)
            industry_high.append(avg_high)
            industry_low.append(avg_low)
            industry_close.append(avg_close)
            industry_volume.append(avg_volume)
            industry_amount.append(avg_amount)

        # 创建行业指数DataFrame
        industry_df = pd.DataFrame({
            'open': industry_open,
            'high': industry_high,
            'low': industry_low,
            'close': industry_close,
            'volume': industry_volume,
            'amount': industry_amount
        }, index=pd.to_datetime(common_dates))

        industry_indices[industry] = industry_df

    return industry_indices


def get_representative_stocks_for_all_industries(industry_mapping, days=200):
    """
    获取所有行业中代表性股票的代码
    """
    industries = industry_mapping['industry'].unique()

    # 存储所有代表性股票
    all_representative_stocks = set()

    for industry in tqdm(industries, desc="分析行业代表性股票", unit="行业"):
        # 获取该行业所有股票
        stocks_in_industry = industry_mapping[industry_mapping['industry'] == industry]['code'].tolist()

        # 临时获取这些股票的数据用于计算代表性
        temp_stock_data_dict = {}
        for stock_code in stocks_in_industry[:50]:  # 限制每个行业最多分析前50只股票，避免过多API调用
            try:
                stock_data = get_stock_data_cached(stock_code, days=min(days, 100))  # 使用较短周期快速评估
                if stock_data is not None and not stock_data.empty:
                    temp_stock_data_dict[stock_code] = stock_data
            except Exception as e:
                print(f"获取股票 {stock_code} 临时数据失败: {e}")

        # 选择该行业的代表性股票
        if temp_stock_data_dict:
            representative_stocks = select_representative_stocks_for_industry(
                list(temp_stock_data_dict.keys()),
                temp_stock_data_dict,
                n=20
            )
            all_representative_stocks.update(representative_stocks)

    return list(all_representative_stocks)


def get_all_industries_and_build_indices():
    """
    获取所有行业并构建行业指数
    """
    print("获取股票行业映射...")
    industry_mapping = get_stock_industry_mapping()

    if industry_mapping is None:
        print("无法获取行业映射，返回空结果")
        return {}, {}  # 返回空字典

    print(f"获取到 {len(industry_mapping)} 条行业映射数据")

    # 获取行业中的股票代码
    industries = industry_mapping['industry'].unique()

    # 构建行业指数
    industry_indices = {}

    # 限制行业数量以减少API请求
    industries_to_process = industries[:30]  # 只处理前30个行业
    print(f"将处理 {len(industries_to_process)} 个行业")

    # 获取代表性股票代码
    print("分析各行业代表性股票...")
    industry_mapping_subset = industry_mapping[industry_mapping['industry'].isin(industries_to_process)]
    representative_stocks = get_representative_stocks_for_all_industries(industry_mapping_subset, days=200)

    print(f"需要获取 {len(representative_stocks)} 只代表性股票的数据")

    # 批量获取代表性股票数据
    stock_data_dict = {}
    for stock_code in tqdm(representative_stocks, desc="获取代表性股票数据", unit="股票"):
        try:
            stock_data = get_stock_data_cached(stock_code, days=200)
            if stock_data is not None and not stock_data.empty:
                stock_data_dict[stock_code] = stock_data
        except Exception as e:
            print(f"获取股票 {stock_code} 数据失败: {e}")

    # 为每个行业计算指数
    for i, industry in enumerate(tqdm(industries_to_process, desc="处理行业", unit="行业")):
        print(f"正在处理行业 {i + 1}/{len(industries_to_process)}: {industry}")

        # 获取该行业的股票代码
        stocks_in_industry = industry_mapping[industry_mapping['industry'] == industry]['code'].tolist()

        # 筛选出已有数据的股票
        valid_stocks = [stock for stock in stocks_in_industry if stock in stock_data_dict]

        if len(valid_stocks) > 0:
            # 获取该行业的股票数据
            industry_stock_data = {stock: stock_data_dict[stock] for stock in valid_stocks}

            # 计算该行业的指数
            industry_data = calculate_industry_index_from_components(
                industry_stock_data,
                industry_mapping[industry_mapping['industry'] == industry][['code', 'industry']]
            )

            if industry in industry_data:
                industry_indices[industry] = industry_data[industry]
                print(f"  - 行业 {industry} 包含 {len(industry_data[industry])} 个交易日数据")
        else:
            print(f"  - 行业 {industry} 没有有效的股票数据")

    return industry_indices, stock_data_dict  # 返回两个字典


def align_data_sequences(df_list, target_length=None):
    """
    对齐数据序列到相同长度
    """
    if not df_list:
        return df_list

    if target_length is None:
        # 使用最短序列的长度作为目标长度
        target_length = min(len(df) for df in df_list)

    aligned_df_list = []
    for df in df_list:
        if len(df) >= target_length:
            # 截取最新的target_length个数据点
            aligned_df = df.tail(target_length).copy()
            aligned_df_list.append(aligned_df)
        else:
            # 如果数据不足，跳过该序列
            print(f"警告: 数据长度 {len(df)} 小于目标长度 {target_length}，跳过该序列")
            continue

    return aligned_df_list


def find_best_params_for_stock(df, pred_len, model, tokenizer):
    """
    为单个股票找到最佳参数组合
    """
    # 定义参数搜索空间
    T_values = [0.5, 0.7, 1.0, 1.3, 1.5]
    top_p_values = [0.7, 0.8, 0.9]
    sample_count = 3

    best_params = None
    best_score = float('inf')  # 最小化误差

    # 创建预测器
    predictor = KronosPredictor(model, tokenizer, max_context=512)

    # 用于验证的最后几天数据
    validation_days = min(5, len(df) // 4)  # 取最后5天或1/4数据用于验证
    if validation_days <= 0:
        return {'T': 1.0, 'top_p': 0.9}  # 返回默认参数

    # 分割训练和验证数据
    train_df = df.iloc[:-validation_days]
    val_df = df.iloc[-validation_days:]

    for T in T_values:
        for top_p in top_p_values:
            try:
                # 使用训练数据预测验证期间
                max_lookback = min(250, len(train_df) - 50)
                lookback = max_lookback

                # 准备输入数据
                x_df = train_df.iloc[-lookback:, :][['open', 'high', 'low', 'close', 'volume', 'amount']].copy()
                x_timestamp = pd.Series(train_df.index[-lookback:])

                # 生成验证期间的时间戳
                val_timestamps = pd.DatetimeIndex(val_df.index)
                y_timestamp_series = pd.Series(val_timestamps)

                # 执行预测
                pred_df = predictor.predict(
                    df=x_df,
                    x_timestamp=x_timestamp,
                    y_timestamp=y_timestamp_series,
                    pred_len=validation_days,
                    T=T,
                    top_p=top_p,
                    sample_count=sample_count,
                    verbose=False
                )

                # 计算预测误差（MAE）
                actual_prices = val_df['close'].values
                predicted_prices = pred_df['close'].values
                mae = np.mean(np.abs(actual_prices - predicted_prices))

                # 更新最佳参数
                if mae < best_score:
                    best_score = mae
                    best_params = {'T': T, 'top_p': top_p}

            except Exception as e:
                print(f"参数组合 (T={T}, top_p={top_p}) 验证失败: {e}")
                continue

    # 如果没找到有效参数，返回默认参数
    if best_params is None:
        best_params = {'T': 1.0, 'top_p': 0.9}

    return best_params


def predict_stock_with_best_params(df, pred_len, model, tokenizer, best_params):
    """
    使用最佳参数预测单个股票
    """
    max_lookback = min(250, len(df) - 50)
    lookback = max_lookback

    # 准备输入数据
    x_df = df.iloc[-lookback:, :][['open', 'high', 'low', 'close', 'volume', 'amount']].copy()
    x_timestamp = pd.Series(df.index[-lookback:])

    # 生成未来交易日时间戳
    future_timestamps = generate_future_trading_days(df.index[-1], pred_len)
    y_timestamp = pd.DatetimeIndex(future_timestamps)
    y_timestamp_series = pd.Series(y_timestamp)

    # 创建预测器
    predictor = KronosPredictor(model, tokenizer, max_context=512)

    # 执行预测
    pred_df = predictor.predict(
        df=x_df,
        x_timestamp=x_timestamp,
        y_timestamp=y_timestamp_series,
        pred_len=pred_len,
        T=best_params['T'],
        top_p=best_params['top_p'],
        sample_count=3,
        verbose=False
    )

    return pred_df


def predict_with_batch_params(df_list, pred_len, model, tokenizer, T=1.0, top_p=0.9, sample_count=1):
    """使用批量预测方法进行多资产预测"""
    # 对齐数据序列
    aligned_df_list = align_data_sequences(df_list)

    if not aligned_df_list:
        print("没有足够的对齐数据进行批量预测")
        return []

    x_df_list = []
    x_timestamp_list = []
    y_timestamp_list = []

    for df in aligned_df_list:
        max_lookback = min(250, len(df) - 50)
        lookback = max_lookback

        # 准备输入数据
        x_df = df.iloc[-lookback:, :][['open', 'high', 'low', 'close', 'volume', 'amount']].copy()
        x_timestamp = pd.Series(df.index[-lookback:])

        # 生成未来交易日时间戳
        future_timestamps = generate_future_trading_days(df.index[-1], pred_len)
        y_timestamp = pd.DatetimeIndex(future_timestamps)
        y_timestamp_series = pd.Series(y_timestamp)

        x_df_list.append(x_df)
        x_timestamp_list.append(x_timestamp)
        y_timestamp_list.append(y_timestamp_series)

    # 创建预测器
    predictor = KronosPredictor(model, tokenizer, max_context=512)

    # 执行批量预测
    pred_df_list = predictor.predict_batch(
        df_list=x_df_list,
        x_timestamp_list=x_timestamp_list,
        y_timestamp_list=y_timestamp_list,
        pred_len=pred_len,
        T=T,
        top_p=top_p,
        sample_count=sample_count,
        verbose=False
    )

    return pred_df_list


def predict_with_ensemble_and_confidence(df_list, pred_len, model, tokenizer,
                                         T_values=[0.7, 1.0, 1.3],
                                         top_p_values=[0.8, 0.9, 0.95],
                                         sample_count=3):
    """使用集成方法和置信度评估进行预测，提高准确性"""
    # 首先对齐数据序列
    aligned_df_list = align_data_sequences(df_list)

    if not aligned_df_list:
        print("没有足够的对齐数据进行集成预测")
        return [], []

    all_pred_results = []

    # 尝试不同的参数组合
    for T in tqdm(T_values, desc="参数T", leave=False):
        for top_p in tqdm(top_p_values, desc="参数top_p", leave=False):
            try:
                pred_df_list = predict_with_batch_params(
                    aligned_df_list,
                    pred_len,
                    model,
                    tokenizer,
                    T=T,
                    top_p=top_p,
                    sample_count=sample_count
                )

                if pred_df_list:  # 确保预测成功
                    all_pred_results.append(pred_df_list)
                else:
                    print(f"警告: 参数组合 (T={T}, top_p={top_p}) 预测失败")

            except Exception as e:
                print(f"参数组合 (T={T}, top_p={top_p}) 预测失败: {e}")
                continue

    if not all_pred_results:
        print("所有参数组合预测都失败了")
        return [], []

    # 对每个板块计算预测值的统计信息
    final_pred_list = []
    confidence_scores = []

    for sector_idx in tqdm(range(len(aligned_df_list)), desc="计算预测结果", unit="行业"):
        # 收集所有参数组合下的预测结果
        sector_predictions = []
        for pred_result in all_pred_results:
            if sector_idx < len(pred_result):  # 确保索引有效
                sector_predictions.append(pred_result[sector_idx])

        if not sector_predictions:
            print(f"行业 {sector_idx} 没有有效的预测结果")
            # 添加空的预测结果以保持索引对齐
            final_pred_list.append(None)
            confidence_scores.append(0.0)
            continue

        # 计算多个预测结果的平均值和置信区间
        valid_predictions = []
        for pred in sector_predictions:
            if pred is not None and not pred.empty:
                valid_predictions.append(pred['close'].iloc[-1])

        if not valid_predictions:
            print(f"行业 {sector_idx} 没有有效的收盘价预测")
            final_pred_list.append(None)
            confidence_scores.append(0.0)
            continue

        close_prices = valid_predictions
        mean_close = np.mean(close_prices)
        std_close = np.std(close_prices)

        # 使用原始数据的最后价格作为基准
        base_price = aligned_df_list[sector_idx]['close'].iloc[-1]

        # 创建预测结果DataFrame（使用第一个有效预测作为模板）
        template_pred = sector_predictions[0]
        final_pred = template_pred.copy()
        final_pred.loc[final_pred.index[-1], 'close'] = mean_close

        final_pred_list.append(final_pred)

        # 计算置信分数（标准差越小，置信度越高）
        if std_close == 0:  # 避免除零错误
            confidence = 1.0
        else:
            confidence = 1 / (1 + std_close / abs(base_price))  # 归一化到0-1
        confidence_scores.append(confidence)

    return final_pred_list, confidence_scores


def predict_industry_growth_from_components(industry_indices, days=10, model=None, tokenizer=None,
                                            use_ensemble=True, confidence_threshold=0.6):
    """
    基于行业指数预测行业未来涨幅
    """
    try:
        # 提取行业名称和对应的数据
        industries = list(industry_indices.keys())
        df_list = [industry_indices[industry] for industry in industries]

        if not df_list:
            print("没有有效的行业数据可供预测")
            return []

        # 根据是否使用集成方法选择预测策略
        if use_ensemble:
            pred_df_list, confidence_scores = predict_with_ensemble_and_confidence(
                df_list, days, model, tokenizer
            )
        else:
            aligned_df_list = align_data_sequences(df_list)
            pred_df_list = predict_with_batch_params(
                aligned_df_list, days, model, tokenizer, T=1.0, top_p=0.9, sample_count=1
            )
            confidence_scores = [1.0] * len(aligned_df_list)  # 如果不使用集成，给默认置信度

        # 计算涨幅和其他指标
        results = []
        # 确保pred_df_list和confidence_scores的长度与industries匹配
        for i, (industry, df) in enumerate(zip(industries, df_list)):
            # 获取对应的预测结果和置信度
            if i < len(pred_df_list) and pred_df_list[i] is not None:
                pred_df = pred_df_list[i]
                conf_score = confidence_scores[i] if i < len(confidence_scores) else 0.0
            else:
                print(f"行业 {industry} 没有有效的预测结果")
                continue

            if conf_score >= confidence_threshold and not pred_df.empty:
                current_price = df['close'].iloc[-1]
                future_price = pred_df['close'].iloc[-1]
                growth_rate = (future_price - current_price) / current_price

                results.append({
                    'industry': industry,
                    'growth_rate': growth_rate,
                    'current_price': current_price,
                    'predicted_price': future_price,
                    'confidence_score': conf_score,
                    'volatility': df['close'].pct_change().std(),  # 历史波动率
                    'volume_trend': df['volume'].tail(10).mean() / df['volume'].head(10).mean()  # 成交量趋势
                })

        return results
    except Exception as e:
        print(f"行业预测失败: {e}")
        import traceback
        traceback.print_exc()
        return []


def save_industry_predictions_to_csv(results, filename="industry_predictions_from_components.csv"):
    """
    将行业预测结果保存到CSV文件
    """
    import csv
    from datetime import datetime

    # 按涨幅排序
    sorted_results = sorted(results, key=lambda x: x['growth_rate'], reverse=True)

    # 添加排名
    for i, result in enumerate(sorted_results, 1):
        result['rank'] = i
        result['prediction_date'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    # 写入CSV
    fieldnames = ['rank', 'industry', 'growth_rate', 'current_price', 'predicted_price',
                  'confidence_score', 'volatility', 'volume_trend', 'prediction_date']
    with open(filename, 'w', newline='', encoding='utf-8-sig') as csvfile:
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
        writer.writeheader()
        for result in sorted_results:
            writer.writerow({
                'rank': result.get('rank', ''),
                'industry': result.get('industry', ''),
                'growth_rate': f"{result.get('growth_rate', 0) * 100:.4f}%",  # 转换为百分比
                'current_price': round(result.get('current_price', 0), 4),
                'predicted_price': round(result.get('predicted_price', 0), 4),
                'confidence_score': round(result.get('confidence_score', 0), 4),
                'volatility': round(result.get('volatility', 0), 6),
                'volume_trend': round(result.get('volume_trend', 0), 4),
                'prediction_date': result.get('prediction_date', '')
            })

    print(f"行业预测结果已保存到 {filename}，共 {len(sorted_results)} 个行业")


def save_detailed_results_to_excel(all_results, industry_indices, industry_mapping, stock_data_dict,
                                   filename="industry_detailed_analysis.xlsx"):
    """
    将详细的预测结果和成分股信息保存到Excel文件
    """

    wb = Workbook()

    # 删除默认工作表
    ws_default = wb.active
    wb.remove(ws_default)

    # 按涨幅排序行业
    sorted_industries = sorted(all_results, key=lambda x: x['growth_rate'], reverse=True)

    # 创建总览工作表
    ws_summary = wb.create_sheet(title="行业总览")
    summary_headers = ['排名', '行业', '预测涨幅', '当前价格', '预测价格', '置信度', '历史波动率', '成交量趋势']
    ws_summary.append(summary_headers)

    for i, result in enumerate(sorted_industries, 1):
        row = [
            i,
            result['industry'],
            f"{result['growth_rate'] * 100:.2f}%",
            result['current_price'],
            result['predicted_price'],
            result['confidence_score'],
            result['volatility'],
            result['volume_trend']
        ]
        ws_summary.append(row)

    # 为表头添加样式
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

    for cell in ws_summary[1]:
        cell.font = header_font
        cell.fill = header_fill

    # 为每个行业创建单独的工作表
    for result in sorted_industries:
        industry_name = result['industry']

        # 获取该行业的成分股信息
        industry_stocks = industry_mapping[industry_mapping['industry'] == industry_name]['code'].tolist()

        # 创建行业工作表
        ws_industry = wb.create_sheet(title=industry_name[:30])  # 限制工作表名称长度

        # 添加行业概览信息
        ws_industry.append(['行业概览'])
        ws_industry.append(['项目', '值'])
        ws_industry.append(['行业名称', industry_name])
        ws_industry.append(['预测涨幅', f"{result['growth_rate'] * 100:.2f}%"])
        ws_industry.append(['当前价格', result['current_price']])
        ws_industry.append(['预测价格', result['predicted_price']])
        ws_industry.append(['置信度', result['confidence_score']])
        ws_industry.append(['历史波动率', result['volatility']])
        ws_industry.append(['成交量趋势', result['volume_trend']])

        # 添加空行
        ws_industry.append([])

        # 添加成分股信息
        ws_industry.append(['成分股详细信息'])

        # 获取成分股数据
        stock_info = []
        for stock_code in industry_stocks:
            if stock_code in stock_data_dict:  # 直接使用传入的stock_data_dict
                stock_df = stock_data_dict[stock_code]
                current_price = stock_df['close'].iloc[-1]
                volatility = stock_df['close'].pct_change().std()
                volume_trend = stock_df['volume'].tail(10).mean() / stock_df['volume'].head(10).mean()

                # 尝试找到最佳参数
                try:
                    best_params = find_best_params_for_stock(stock_df, 10, globals()['model'], globals()['tokenizer'])
                    stock_info.append({
                        'code': stock_code,
                        'current_price': current_price,
                        'volatility': volatility,
                        'volume_trend': volume_trend,
                        'best_T': best_params['T'],
                        'best_top_p': best_params['top_p']
                    })
                except:
                    stock_info.append({
                        'code': stock_code,
                        'current_price': current_price,
                        'volatility': volatility,
                        'volume_trend': volume_trend,
                        'best_T': 'N/A',
                        'best_top_p': 'N/A'
                    })

        # 添加成分股表头
        ws_industry.append(['股票代码', '当前价格', '历史波动率', '成交量趋势', '最佳T参数', '最佳top_p参数'])

        # 添加成分股数据
        for info in stock_info:
            ws_industry.append([
                info['code'],
                info['current_price'],
                info['volatility'],
                info['volume_trend'],
                info['best_T'],
                info['best_top_p']
            ])

    # 保存Excel文件
    wb.save(filename)
    print(f"详细分析结果已保存到 {filename}")


def analyze_industries_from_components(prediction_days=10,
                                       output_filename="industry_predictions_from_components.csv"):
    """
    基于成分股分析行业并保存预测结果
    """
    print(f"开始基于成分股分析行业未来{prediction_days}天的预测涨幅...")

    # 加载模型和分词器
    print("正在加载模型和分词器...")
    try:
        tokenizer = KronosTokenizer.from_pretrained("NeoQuasar/Kronos-Tokenizer-base")
        model = Kronos.from_pretrained("NeoQuasar/Kronos-small")
        print("模型加载完成")
    except Exception as e:
        print(f"模型加载失败: {e}")
        return

    # 保存模型和分词器到全局变量以便在其他函数中使用
    globals()['model'] = model
    globals()['tokenizer'] = tokenizer

    # 获取行业映射
    print("获取行业映射...")
    industry_mapping = get_stock_industry_mapping()

    # 构建行业指数
    print("正在构建行业指数...")
    industry_indices, stock_data_dict = get_all_industries_and_build_indices()

    if not industry_indices:
        print("没有构建出任何行业指数")
        return

    print(f"共构建了 {len(industry_indices)} 个行业指数")

    # 预测行业涨幅
    print("正在预测行业涨幅...")
    all_results = predict_industry_growth_from_components(
        industry_indices, prediction_days, model, tokenizer
    )

    if not all_results:
        print("没有成功预测任何行业")
        return

    # 保存结果到CSV
    save_industry_predictions_to_csv(all_results, output_filename)

    # 保存详细结果到Excel
    excel_filename = output_filename.replace('.csv', '_detailed.xlsx')
    save_detailed_results_to_excel(all_results, industry_indices, industry_mapping, stock_data_dict, excel_filename)

    # 显示前10名
    top_10 = sorted(all_results, key=lambda x: x['growth_rate'], reverse=True)[:10]
    print("\n=== 预计未来10天涨幅前10的行业 ===")
    for i, item in enumerate(top_10, 1):
        print(f"{i}. {item['industry']} - 预计涨幅: {item['growth_rate'] * 100:.2f}% "
              f"(置信度: {item['confidence_score']:.3f})")

    # 显示后10名（跌幅最大）
    bottom_10 = sorted(all_results, key=lambda x: x['growth_rate'])[:10]
    print("\n=== 预计未来10天跌幅前10的行业 ===")
    for i, item in enumerate(bottom_10, 1):
        print(f"{i}. {item['industry']} - 预计跌幅: {item['growth_rate'] * 100:.2f}% "
              f"(置信度: {item['confidence_score']:.3f})")

    return all_results


def predict_top_growing_industries():
    """
    预测并返回涨幅前5的行业
    """
    all_results = analyze_industries_from_components(
        prediction_days=10,
        output_filename="industry_predictions_from_components_" + datetime.now().strftime('%Y%m%d_%H%M%S') + ".csv"
    )

    if all_results:
        # 获取前5名
        top_5 = sorted(all_results, key=lambda x: x['growth_rate'], reverse=True)[:5]
        return top_5
    else:
        return []


if __name__ == "__main__":
    # 执行基于成分股的行业预测
    all_results = analyze_industries_from_components(
        prediction_days=10,
        output_filename="industry_predictions_from_components_" + datetime.now().strftime('%Y%m%d_%H%M%S') + ".csv"
    )

    if all_results:
        # 获取前5名
        top_5 = sorted(all_results, key=lambda x: x['growth_rate'], reverse=True)[:5]
        print("\n预测完成！以下是涨幅前5的行业：")
        for i, industry in enumerate(top_5, 1):
            print(f"{i}. {industry['industry']} - 预计涨幅: {industry['growth_rate'] * 100:.2f}% "
                  f"(置信度: {industry['confidence_score']:.3f})")
    else:
        print("预测失败，没有获得任何结果")
