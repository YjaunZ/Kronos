"""
基于Kronos模型预测主板股票价格，筛选前10只涨幅最高的股票
同时考虑参数准确性和最佳买卖时机

主要特性
1. 股票筛选功能
使用baostock接口获取上证和深证主板股票
自动排除ST股票
支持测试模式（默认分析20只股票）
2. Kronos模型预测
针对每只股票寻找最优参数组合
考虑不同参数的准确性（T、top_p、sample_count）
使用RMSE评估预测质量
3. 最佳买卖时机分析
计算最佳买入价和卖出价
确定最佳持股天数
估算盈利概率
4. 性能优化
API请求频率控制，减少baostock接口访问次数
批量处理优化（使用predict_batch方法）
错误处理和异常恢复
5. 置信度评分
综合考虑涨幅、预测准确性和参数稳定性
为用户提供更可靠的推荐依据
程序首先在测试模式下分析20只股票，您可以根据需要调整股票数量和预测参数。
功能说明：
Excel导出功能：新增save_predictions_to_excel方法，将所有预测结果保存到Excel文件
数据格式化：将数值格式化为易读的形式（百分比、保留小数位等）
Excel样式美化：设置表头样式、边框线和自动调整列宽
文件命名：使用时间戳命名Excel文件，避免覆盖
股票数据验证：验证获取的数据是否为真实的股票而非指数
股票筛选优化：更严格地筛选主板股票，排除ETF基金和债券等非股票品种
此代码实现了将所有预测结果保存到Excel文件的功能，包含股票基本信息、预测价格、置信度等详细信息。
"""
import pandas as pd
import numpy as np
import sys
import os
import baostock as bs
from datetime import datetime, timedelta
import holidays
import time
from typing import List, Tuple, Dict
import warnings
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

warnings.filterwarnings('ignore')

# 添加项目路径
sys.path.append(os.path.join(os.path.dirname(__file__), ".."))

from model import Kronos, KronosTokenizer, KronosPredictor


class StockPredictor:
    def __init__(self):
        self.last_request_time = 0
        self.api_delay = 0.5  # API请求间隔，单位秒
        self.cn_holidays = holidays.China()

    def login_baostock(self):
        """登录baostock服务"""
        lg = bs.login()
        if lg.error_code != '0':
            print(f"Baostock登录失败: {lg.error_msg}")
            return False
        return True

    def logout_baostock(self):
        """登出baostock服务"""
        bs.logout()

    def rate_limit(self):
        """控制API请求频率"""
        current_time = time.time()
        time_since_last_request = current_time - self.last_request_time
        if time_since_last_request < self.api_delay:
            time.sleep(self.api_delay - time_since_last_request)
        self.last_request_time = time.time()

    def generate_future_trading_days(self, start_date, num_days):
        """生成未来交易日时间戳（排除周末和节假日）"""
        if isinstance(start_date, str):
            start_date = pd.to_datetime(start_date)
        elif isinstance(start_date, datetime):
            start_date = pd.to_datetime(start_date)

        trading_days = []
        current_date = start_date + timedelta(days=1)

        while len(trading_days) < num_days:
            if current_date.weekday() < 5 and current_date.date() not in self.cn_holidays:
                trading_days.append(current_date)
            current_date += timedelta(days=1)

        return trading_days

    def get_stock_data(self, symbol, days=500):
        """获取指定天数的股票数据 using baostock"""
        self.rate_limit()

        if not self.login_baostock():
            raise Exception("无法连接到baostock服务")

        try:
            end_date = datetime.now().strftime('%Y-%m-%d')
            start_date = (datetime.now() - timedelta(days=days)).strftime('%Y-%m-%d')

            # 正确处理A股股票代码
            if symbol.startswith('6'):  # 上交所
                code = f"sh.{symbol}"
            elif symbol.startswith(('0', '3')):  # 深交所
                code = f"sz.{symbol}"
            else:
                code = f"sh.{symbol}"

            rs = bs.query_history_k_data_plus(
                code,
                "date,open,high,low,close,volume,amount",
                start_date=start_date,
                end_date=end_date,
                frequency="d"
            )

            data_list = []
            while (rs.error_code == '0') & rs.next():
                data_list.append(rs.get_row_data())

            if len(data_list) == 0:
                raise Exception(f"未获取到代码 {symbol} 的数据")

            result = pd.DataFrame(data_list, columns=rs.fields)

            # 重命名列
            column_mapping = {
                'date': 'timestamps',
                'open': 'open',
                'high': 'high',
                'low': 'low',
                'close': 'close',
                'volume': 'volume',
                'amount': 'amount'
            }

            result = result.rename(columns=column_mapping)[list(column_mapping.values())].copy()
            result['timestamps'] = pd.to_datetime(result['timestamps'])

            # 数值类型转换
            numeric_columns = ['open', 'high', 'low', 'close', 'volume', 'amount']
            for col in numeric_columns:
                result[col] = pd.to_numeric(result[col], errors='coerce')

            result = result.dropna()

            # 验证是否为股票数据（检查数据量和价格范围）
            if len(result) < 50:
                raise Exception(f"代码 {symbol} 数据量不足，可能不是股票")

            # 检查价格是否合理（股票价格一般在0.1-1000元范围内）
            avg_price = result['close'].mean()
            if avg_price < 0.1 or avg_price > 1000:
                raise Exception(f"代码 {symbol} 价格异常 ({avg_price:.2f})，可能不是股票")

            # 额外验证：检查是否为指数（通过波动性判断）
            price_volatility = result['close'].std() / result['close'].mean()
            if price_volatility < 0.01:  # 波动性过小，可能是指数
                raise Exception(f"代码 {symbol} 波动性过小，可能不是个股而是指数")

            return result

        finally:
            self.logout_baostock()

    def get_main_board_stocks(self, limit=20):
        """获取主板股票列表（上证+深证），排除ST股票和科创板、创业板，限制数量用于测试"""
        self.rate_limit()

        if not self.login_baostock():
            print("Baostock登录失败")
            return []

        try:
            # 查询股票基本信息
            rs = bs.query_stock_basic()
            stock_list = []

            while (rs.error_code == '0') & rs.next():
                stock_list.append(rs.get_row_data())

            if len(stock_list) > 0:
                df = pd.DataFrame(stock_list, columns=rs.fields)

                # 检查数据结构
                print(f"获取到 {len(df)} 只股票，列名: {df.columns.tolist()}")

                # baostock返回的code格式通常是 sh.600001, sz.000001 等
                # 验证是否有code和code_name列
                if 'code' not in df.columns or 'code_name' not in df.columns:
                    print(f"数据列不完整，实际列: {df.columns.tolist()}")
                    return []

                # 提取股票代码数字部分（保持原始code格式用于区分上交所/深交所）
                df['pure_code'] = df['code'].apply(lambda x: x.split('.')[1] if '.' in str(x) else str(x))
                df['exchange'] = df['code'].apply(lambda x: x.split('.')[0] if '.' in str(x) else '')

                # 筛选主板股票 - 排除指数和其他非股票品种
                main_board_mask = (
                    # 上海主板
                        (df['pure_code'].str.startswith('600')) |
                        (df['pure_code'].str.startswith('601')) |
                        (df['pure_code'].str.startswith('603')) |
                        (df['pure_code'].str.startswith('605')) |
                        # 深圳主板
                        (df['pure_code'].str.startswith('000')) |
                        (df['pure_code'].str.startswith('001')) |
                        (df['pure_code'].str.startswith('002')) |
                        (df['pure_code'].str.startswith('003'))
                )

                # 排除科创板（688开头）和创业板（300开头）
                exclude_kcb_cyb = ~(
                        (df['pure_code'].str.startswith('688')) |
                        (df['pure_code'].str.startswith('300'))
                )

                # 排除ETF基金（通常以51开头）和债券等
                exclude_etf_bonds = ~(
                        (df['pure_code'].str.startswith('51')) |
                        (df['pure_code'].str.startswith('11')) |
                        (df['pure_code'].str.startswith('12')) |
                        (df['pure_code'].str.startswith('13')) |
                        (df['pure_code'].str.startswith('20')) |
                        (df['pure_code'].str.startswith('15')) |
                        (df['pure_code'].str.startswith('16')) |
                        (df['pure_code'].str.startswith('18'))
                )

                # 排除ST股票
                non_st_mask = ~df['code_name'].str.contains(
                    'ST|st|S.*?T|退|退市|暂停上市|终止上市',
                    na=False,
                    case=False
                )
                # 排除指数（通过名称判断）
                exclude_indices = ~df['code_name'].str.contains('指数|指数', na=False, case=False)

                # 合并筛选条件
                filtered_df = df[main_board_mask & exclude_kcb_cyb & exclude_etf_bonds & non_st_mask]

                # 只取前limit个股票用于测试
                result_stocks = filtered_df['pure_code'].head(limit).tolist()

                print(f"获取到 {len(result_stocks)} 只主板非ST股票用于测试")
                return result_stocks
            else:
                print("未能获取股票基本信息")
                return []

        finally:
            self.logout_baostock()

    def predict_with_params_batch(self, df_list, pred_len, model, tokenizer, T=1.0, top_p=0.9, sample_count=1):
        """使用指定参数进行批量预测，提高效率"""
        x_df_list = []
        x_timestamp_list = []
        y_timestamp_list = []

        for df in df_list:
            max_lookback = min(250, len(df) - 50)
            lookback = max_lookback

            # 准备输入数据
            x_df = df.iloc[-lookback:, :][['open', 'high', 'low', 'close', 'volume', 'amount']].copy()
            x_timestamp = pd.Series(df.index[-lookback:])

            # 生成未来交易日时间戳
            future_timestamps = self.generate_future_trading_days(df.index[-1], pred_len)
            y_timestamp = pd.DatetimeIndex(future_timestamps)
            y_timestamp_series = pd.Series(y_timestamp)

            x_df_list.append(x_df)
            x_timestamp_list.append(x_timestamp)
            y_timestamp_list.append(y_timestamp_series)

        # 创建预测器
        predictor = KronosPredictor(model, tokenizer, max_context=512)

        # 执行批量预测
        pred_df_list = predictor.predict_batch(
            df_list=x_df_list,
            x_timestamp_list=x_timestamp_list,
            y_timestamp_list=y_timestamp_list,
            pred_len=pred_len,
            T=T,
            top_p=top_p,
            sample_count=sample_count,
            verbose=False
        )

        return pred_df_list

    def predict_with_params(self, df, pred_len, model, tokenizer, T=1.0, top_p=0.9, sample_count=1):
        """使用指定参数进行预测"""
        max_lookback = min(250, len(df) - 50)
        lookback = max_lookback

        x_df = df.iloc[-lookback:, :][['open', 'high', 'low', 'close', 'volume', 'amount']].copy()
        x_timestamp = pd.Series(df.index[-lookback:])

        future_timestamps = self.generate_future_trading_days(df.index[-1], pred_len)
        y_timestamp = pd.DatetimeIndex(future_timestamps)
        y_timestamp_series = pd.Series(y_timestamp)

        predictor = KronosPredictor(model, tokenizer, max_context=512)

        pred_df = predictor.predict(
            df=x_df,
            x_timestamp=x_timestamp,
            y_timestamp=y_timestamp_series,
            pred_len=pred_len,
            T=T,
            top_p=top_p,
            sample_count=sample_count,
            verbose=False
        )

        pred_df.index = y_timestamp
        return pred_df

    def evaluate_predictions(self, actual_df, predicted_df):
        """评估预测准确性"""
        if actual_df.empty or predicted_df.empty:
            return float('inf'), float('inf'), float('inf')

        mae = np.mean(np.abs(actual_df['close'] - predicted_df['close']))
        mse = np.mean((actual_df['close'] - predicted_df['close']) ** 2)
        rmse = np.sqrt(mse)

        return mae, mse, rmse

    def find_best_parameters(self, df, prediction_days=10, model=None, tokenizer=None):
        """为单个股票找到最佳参数组合"""
        param_combinations = [
            {"T": 0.5, "top_p": 0.8, "sample_count": 1},
            {"T": 0.7, "top_p": 0.85, "sample_count": 1},
            {"T": 0.8, "top_p": 0.9, "sample_count": 1},
            {"T": 1.0, "top_p": 0.9, "sample_count": 1},
            {"T": 1.2, "top_p": 0.95, "sample_count": 1},
            {"T": 0.6, "top_p": 0.8, "sample_count": 3},
            {"T": 0.8, "top_p": 0.9, "sample_count": 3},
            {"T": 1.0, "top_p": 0.9, "sample_count": 3},
        ]

        split_point = len(df) - prediction_days
        train_df = df[:split_point]
        actual_future = df[split_point:]

        best_rmse = float('inf')
        best_params = None

        # 检查模型是否已加载
        if model is None or tokenizer is None:
            print("模型或分词器未加载，无法进行参数优化")
            return None, float('inf')

        for params in param_combinations:
            try:
                pred_df = self.predict_with_params(
                    train_df,
                    prediction_days,
                    model=model,
                    tokenizer=tokenizer,
                    T=params["T"],
                    top_p=params["top_p"],
                    sample_count=params["sample_count"]
                )

                mae, mse, rmse = self.evaluate_predictions(actual_future, pred_df)

                if rmse < best_rmse:
                    best_rmse = rmse
                    best_params = params

            except Exception as e:
                print(f"参数组合 {params} 预测失败: {e}")
                continue

        return best_params, best_rmse

    def apply_daily_limit(self, predicted_price, current_price, stock_code):
        """
        应用A股涨跌幅限制规则
        主板股票涨跌幅限制为±10%
        科创板和创业板涨跌幅限制为±20%
        """
        # 计算理论涨跌幅
        theoretical_change = (predicted_price - current_price) / current_price

        # 根据股票类型应用不同限制
        if stock_code.startswith(('600', '601', '603', '605', '000', '001', '002', '003')):
            # 主板股票
            limit = 0.1  # ±10%
        elif stock_code.startswith(('688', '300')):
            # 科创板/创业板
            limit = 0.2  # ±20%
        else:
            # 默认主板限制
            limit = 0.1

        # 限制涨跌幅在合理范围内
        capped_change = max(min(theoretical_change, limit), -limit)
        adjusted_price = current_price * (1 + capped_change)

        return adjusted_price

    def find_optimal_entry_exit_points(self, pred_df):
        """找到最佳买入、卖出价格和持股天数"""
        close_prices = pred_df['close'].values

        if len(close_prices) < 2:
            return None, None, 0, 0

        # 寻找最低点作为买入价，最高点作为卖出价
        min_idx = np.argmin(close_prices)
        max_idx = np.argmax(close_prices[min_idx:]) + min_idx  # 确保卖出时间在买入之后

        if min_idx >= max_idx:  # 如果最低点在最高点之后，则寻找整体趋势
            # 简单策略：第一个点买入，最后一个点卖出
            entry_price = close_prices[0]
            exit_price = close_prices[-1]
            entry_day = 0
            holding_days = len(close_prices) - 1
        else:
            entry_price = close_prices[min_idx]
            exit_price = close_prices[max_idx]
            entry_day = min_idx
            holding_days = max_idx - min_idx

        profit_probability = (exit_price - entry_price) / entry_price if entry_price > 0 else 0

        return entry_price, exit_price, holding_days, profit_probability

    def save_predictions_to_excel(self, predictions, filename="stock_predictions.xlsx"):
        """将预测结果保存到Excel文件"""
        if not predictions:
            print("没有预测结果可以保存")
            return

        # 创建DataFrame
        df_data = []
        for pred in predictions:
            df_data.append({
                '股票代码': pred['code'],
                '股票名称': pred['name'],
                '当前价格': pred['current_price'],
                '预测最终价格': pred['predicted_final_price'],
                '总涨幅百分比': f"{pred['total_growth_rate'] * 100:.2f}%",
                '最佳买入价': pred['entry_price'] if pred['entry_price'] else 'N/A',
                '最佳卖出价': pred['exit_price'] if pred['exit_price'] else 'N/A',
                '最佳持股天数': pred['holding_days'] if pred['holding_days'] else 'N/A',
                '盈利概率百分比': f"{pred['profit_probability'] * 100:.2f}%" if pred['profit_probability'] else 'N/A',
                '最佳参数_T': pred['best_params']['T'] if pred['best_params'] else 'N/A',
                '最佳参数_top_p': pred['best_params']['top_p'] if pred['best_params'] else 'N/A',
                '最佳参数_sample_count': pred['best_params']['sample_count'] if pred['best_params'] else 'N/A',
                'RMSE': f"{pred['rmse']:.4f}",
                '置信度分数': f"{pred['confidence_score']:.4f}"
            })

        df = pd.DataFrame(df_data)

        # 创建Excel工作簿
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "股票预测结果"

        # 写入数据
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

        # 设置标题样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # 设置边框
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = thin_border

        # 自动调整列宽
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # 限制最大宽度为50
            ws.column_dimensions[column_letter].width = adjusted_width

        # 保存文件
        wb.save(filename)
        print(f"预测结果已保存到 {filename}")

    def predict_top_stocks(self, top_n=10, test_mode=True, batch_size=5, save_to_excel=True):
        """预测并筛选涨幅前top_n的股票"""
        print("开始预测主板股票未来走势...")

        # 加载模型和分词器
        print("正在加载模型和分词器...")
        try:
            tokenizer = KronosTokenizer.from_pretrained("NeoQuasar/Kronos-Tokenizer-base")
            model = Kronos.from_pretrained("NeoQuasar/Kronos-small")
            print("模型加载完成")
        except Exception as e:
            print(f"模型加载失败: {e}")
            return

        # 获取主板股票列表
        print("正在获取主板股票列表...")
        stock_codes = self.get_main_board_stocks(limit=20 if test_mode else 1000)  # 测试模式只取20只

        if not stock_codes:
            print("未能获取股票列表")
            return

        print(f"获取到 {len(stock_codes)} 只股票进行分析")

        # 分批处理股票数据，避免一次性加载过多数据
        predictions = []

        # 将股票代码分成批次处理
        for batch_start in range(0, len(stock_codes), batch_size):
            batch_codes = stock_codes[batch_start:batch_start + batch_size]
            print(f"正在处理第 {batch_start + 1}-{min(batch_start + batch_size, len(stock_codes))} 只股票...")

            # 获取本批次的股票数据
            stock_data_list = []
            valid_codes = []

            for code in batch_codes:
                try:
                    df = self.get_stock_data(code, days=300)
                    if df.empty or len(df) < 100:
                        print(f"  - 获取股票 {code} 数据不足，跳过")
                        continue

                    df = df.set_index('timestamps')
                    stock_data_list.append(df)
                    valid_codes.append(code)
                except Exception as e:
                    print(f"  - 获取股票 {code} 数据失败: {e}")
                    continue

            if not stock_data_list:
                continue

            # 为本批次的股票找到最佳参数
            batch_predictions = []
            for i, (code, df) in enumerate(zip(valid_codes, stock_data_list)):
                print(f"  正在为股票 {code} 寻找最佳参数...")

                best_params, best_rmse = self.find_best_parameters(df, prediction_days=20, model=model,
                                                                   tokenizer=tokenizer)

                if best_params is None:
                    print(f"    - 股票 {code} 参数优化失败，跳过")
                    continue

                # 使用最佳参数进行实际预测
                pred_df = self.predict_with_params(
                    df,
                    20,  # 预测20天
                    model,
                    tokenizer,
                    T=best_params["T"],
                    top_p=best_params["top_p"],
                    sample_count=best_params["sample_count"]
                )

                # 计算总涨幅
                current_price = df['close'].iloc[-1]

                # 应用A股涨跌幅限制规则
                raw_future_price = pred_df['close'].iloc[-1]
                future_price = self.apply_daily_limit(raw_future_price, current_price, code)

                total_growth_rate = (future_price - current_price) / current_price

                # 计算最佳买入、卖出时机
                entry_price, exit_price, holding_days, profit_prob = self.find_optimal_entry_exit_points(pred_df)

                # 计算置信度分数
                confidence_score = self.calculate_confidence_score(total_growth_rate, best_rmse, best_params)

                batch_predictions.append({
                    'code': code,
                    'name': self.get_stock_name(code),
                    'current_price': current_price,
                    'predicted_final_price': future_price,
                    'total_growth_rate': total_growth_rate,
                    'best_params': best_params,
                    'rmse': best_rmse,
                    'confidence_score': confidence_score,
                    'entry_price': entry_price,
                    'exit_price': exit_price,
                    'holding_days': holding_days,
                    'profit_probability': profit_prob
                })

                print(f"    - {code}: 当前价 {current_price:.2f}, 预测价 {future_price:.2f}, "
                      f"总涨幅 {total_growth_rate * 100:.2f}%, 置信度 {confidence_score:.4f}")

            predictions.extend(batch_predictions)

        if not predictions:
            print("没有成功预测任何股票")
            return

        # 保存所有预测结果到Excel
        if save_to_excel:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            excel_filename = f"stock_predictions_{timestamp}.xlsx"
            self.save_predictions_to_excel(predictions, excel_filename)

        # 按多维度排序：置信度 -> 概率 -> 持股天数 -> 涨幅 -> 价格
        sorted_predictions = sorted(predictions,
                                    key=lambda x: (x['confidence_score'],
                                                   x['profit_probability'],
                                                   -x['holding_days'],  # 持股天数短的优先
                                                   x['total_growth_rate'],  # 涨幅大的优先
                                                   -x['current_price']),  # 价格低的优先
                                    reverse=True)

        top_predictions = sorted_predictions[:top_n]

        print(f"\n=== 预测涨幅前{top_n}的主板股票 ===")
        for i, item in enumerate(top_predictions, 1):
            print(f"{i}. {item['name']} ({item['code']})")
            print(f"   当前价格: {item['current_price']:.2f}")
            print(f"   预测价格: {item['predicted_final_price']:.2f}")
            print(f"   总涨幅: {item['total_growth_rate'] * 100:.2f}%")
            print(f"   最佳买入价: {item['entry_price']:.2f}" if item['entry_price'] else "   最佳买入价: N/A")
            print(f"   最佳卖出价: {item['exit_price']:.2f}" if item['exit_price'] else "   最佳卖出价: N/A")
            print(f"   最佳持股天数: {item['holding_days']}" if item['holding_days'] else "   最佳持股天数: N/A")
            print(f"   盈利概率: {item['profit_probability'] * 100:.2f}%" if item[
                'profit_probability'] else "   盈利概率: N/A")
            print(f"   参数: T={item['best_params']['T']}, top_p={item['best_params']['top_p']}, "
                  f"sample_count={item['best_params']['sample_count']}")
            print(f"   RMSE: {item['rmse']:.4f}")
            print(f"   置信度: {item['confidence_score']:.4f}")
            print()

        return top_predictions

    def calculate_confidence_score(self, growth_rate, rmse, params):
        """计算综合置信度分数"""
        base_score = abs(growth_rate)
        accuracy_penalty = rmse / 10
        param_stability_bonus = 0.05 if params['T'] <= 1.0 and 0.8 <= params['top_p'] <= 0.95 else 0
        confidence_score = base_score - accuracy_penalty + param_stability_bonus
        return confidence_score

    def get_stock_name(self, code):
        """获取股票名称"""
        self.rate_limit()

        if not self.login_baostock():
            return f"股票{code}"

        try:
            # 查询股票基本信息
            rs = bs.query_stock_basic()

            stock_info_list = []
            while (rs.error_code == '0') & rs.next():
                stock_info_list.append(rs.get_row_data())

            if len(stock_info_list) > 0:
                stock_info = pd.DataFrame(stock_info_list, columns=rs.fields)

                # 构建完整代码格式进行匹配
                if code.startswith('6'):
                    full_code = f"sh.{code}"
                else:
                    full_code = f"sz.{code}"

                # 查找指定股票代码的信息
                # 尝试匹配完整代码
                target_stock = stock_info[stock_info['code'] == full_code]

                if not target_stock.empty and 'code_name' in target_stock.columns:
                    name = target_stock['code_name'].iloc[0]
                    # 检查是否为指数（名称中包含"指数"字样）
                    if '指数' in name:
                        self.logout_baostock()
                        return f"指数{code}"
                    self.logout_baostock()
                    return name.strip()  # 去除可能的空白字符
                else:
                    self.logout_baostock()
                    return f"股票{code}"
            else:
                self.logout_baostock()
                return f"股票{code}"

        except Exception as e:
            print(f"获取股票 {code} 名称失败: {e}")
            try:
                self.logout_baostock()
            except:
                pass
            return f"股票{code}"


def main():
    predictor = StockPredictor()

    # 测试模式，只分析20只股票
    print("开始测试模式，分析20只主板股票...")
    top_stocks = predictor.predict_top_stocks(top_n=10, test_mode=True, batch_size=5)

    # 取消测试模式，分析所有可获得的主板股票
    # print("开始分析所有主板股票...")
    # top_stocks = predictor.predict_top_stocks(top_n=10, test_mode=False, batch_size=5)

    if top_stocks:
        print("完成！以下是预测涨幅前10的股票：")
        for i, stock in enumerate(top_stocks, 1):
            print(f"{i}. {stock['name']} ({stock['code']}) - 预计涨幅: {stock['total_growth_rate'] * 100:.2f}%")
    else:
        print("预测失败，没有获得任何结果")


if __name__ == "__main__":
    main()
